"""
TJK Excel / Google Sheets Rapor Modülü
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os

# ─── Renk Paleti ────────────────────────────────────────────────
RENK = {
    "baslik_bg":   "1A3A5C",   # Koyu lacivert
    "baslik_yazi": "FFFFFF",   # Beyaz
    "altyazi_bg":  "2E6DA4",   # Mavi
    "altyazi_yazi":"FFFFFF",
    "satir_tek":   "EEF4FB",   # Açık mavi
    "satir_cift":  "FFFFFF",   # Beyaz
    "birinci":     "D4EDDA",   # Yeşil (1. derece)
    "ikinci":      "FFF3CD",   # Sarı (2. derece)
    "ucuncu":      "FFE5D0",   # Turuncu (3. derece)
    "ozet_bg":     "F0F8E8",   # Açık yeşil
    "toplam_bg":   "1A3A5C",   # Koyu
    "toplam_yazi": "FFFFFF",
    "komisyon_bg": "E8F5E9",
    "para_format": '#,##0.00 [$₺-41F]',
}

def _ince_cerceve():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _kalin_cerceve():
    medium = Side(style="medium", color="1A3A5C")
    return Border(left=medium, right=medium, top=medium, bottom=medium)

def excel_olustur(
    tum_sonuclar: dict,   # {sahip_isim: [sonuc_listesi]}
    tarih: str,
    komisyon_yuzdesi: float,
    cikti_yolu: str = None
) -> str:
    """
    tum_sonuclar = {
      "AHMET YILMAZ": [ {...}, {...} ],
      "MEHMET DEMİR": [ {...} ],
    }
    Her at sahibi için ayrı sheet + özet sheet oluşturur.
    Döner: kayıt edilen dosya yolu.
    """
    gun_str = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")

    if cikti_yolu is None:
        cikti_yolu = f"TJK_Rapor_{tarih.replace('-','')}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # Varsayılan boş sheet'i sil

    # ── Her sahip için ayrı sheet ──────────────────────────────
    ozet_satirlari = []   # Özet sheet için

    for sahip_isim, sonuclar in tum_sonuclar.items():
        sheet_isim = sahip_isim[:28].strip()   # max 31 char
        ws = wb.create_sheet(title=sheet_isim)
        _sahip_sheet_doldur(ws, sahip_isim, tarih, gun_str,
                             sonuclar, komisyon_yuzdesi)

        # Özet için kayıt
        toplam_ik = sum(s["IkramiyeTL"] for s in sonuclar)
        birinci   = sum(1 for s in sonuclar if str(s["Derece"]).strip()=="1")
        ozet_satirlari.append({
            "sahip":    sahip_isim,
            "at_sayisi": len(sonuclar),
            "birinci":  birinci,
            "toplam":   toplam_ik,
            "komisyon": toplam_ik * komisyon_yuzdesi / 100,
        })

    # ── Özet sheet ────────────────────────────────────────────
    if ozet_satirlari:
        _ozet_sheet_olustur(wb, ozet_satirlari, gun_str, komisyon_yuzdesi)

    wb.save(cikti_yolu)
    return cikti_yolu


def _sahip_sheet_doldur(ws, sahip_isim, tarih, gun_str,
                         sonuclar, komisyon_yuzdesi):
    # ── Başlık ──────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = f"🏇 TJK Koşu Sonuçları  —  {gun_str}"
    ws["A1"].font   = Font(bold=True, size=14, color=RENK["baslik_yazi"],
                           name="Arial")
    ws["A1"].fill   = PatternFill("solid", fgColor=RENK["baslik_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = f"At Sahibi: {sahip_isim}"
    ws["A2"].font   = Font(bold=True, size=11, color=RENK["altyazi_yazi"],
                           name="Arial")
    ws["A2"].fill   = PatternFill("solid", fgColor=RENK["altyazi_bg"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Tablo Başlıkları ─────────────────────────────────────
    KOLONLAR = [
        ("Koşu No",    8),
        ("At Adı",    22),
        ("Hipodrom",  12),
        ("Mesafe",     8),
        ("Zemin",      8),
        ("Grup",       8),
        ("Antrenör",  20),
        ("Derece",     8),
        ("İkramiye (₺)", 16),
    ]

    for col_idx, (baslik, genislik) in enumerate(KOLONLAR, start=1):
        hucre = ws.cell(row=3, column=col_idx, value=baslik)
        hucre.font      = Font(bold=True, color=RENK["baslik_yazi"],
                               name="Arial", size=10)
        hucre.fill      = PatternFill("solid", fgColor=RENK["altyazi_bg"])
        hucre.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        hucre.border    = _ince_cerceve()
        ws.column_dimensions[get_column_letter(col_idx)].width = genislik
    ws.row_dimensions[3].height = 20

    # ── Veri Satırları ───────────────────────────────────────
    for row_idx, s in enumerate(sonuclar, start=4):
        derece_str = str(s.get("Derece", "-")).strip()

        # Satır rengi dereceye göre
        if   derece_str == "1": bg = RENK["birinci"]
        elif derece_str == "2": bg = RENK["ikinci"]
        elif derece_str == "3": bg = RENK["ucuncu"]
        elif row_idx % 2 == 0:  bg = RENK["satir_tek"]
        else:                   bg = RENK["satir_cift"]

        fill = PatternFill("solid", fgColor=bg)

        deger = [
            s.get("KosuNo",   "-"),
            s.get("AtAdi",    "-"),
            s.get("HipodromAdi", "-"),
            s.get("MesafeAdi","-"),
            s.get("ZeminAdi", "-"),
            s.get("GrupAdi",  "-"),
            s.get("AntrenorAdi", "-"),
            derece_str,
            s.get("IkramiyeTL", 0.0),
        ]

        for col_idx, val in enumerate(deger, start=1):
            hucre = ws.cell(row=row_idx, column=col_idx, value=val)
            hucre.fill      = fill
            hucre.border    = _ince_cerceve()
            hucre.font      = Font(name="Arial", size=10)
            hucre.alignment = Alignment(vertical="center")
            if col_idx == 9 and isinstance(val, (int, float)):
                hucre.number_format = RENK["para_format"]
                hucre.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (1, 8):
                hucre.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    # ── Özet Bölümü ─────────────────────────────────────────
    son_satir = 3 + len(sonuclar)
    bos = son_satir + 1

    ik_col_harfi = get_column_letter(9)
    ik_aralik = f"{ik_col_harfi}4:{ik_col_harfi}{son_satir}" \
                if sonuclar else f"{ik_col_harfi}4:{ik_col_harfi}4"

    ozet_veriler = [
        ("Toplam At Sayısı",       len(sonuclar),   None),
        ("Birinci Sayısı",         sum(1 for s in sonuclar if str(s.get("Derece","")).strip()=="1"), None),
        ("Toplam İkramiye (₺)",    f"=SUM({ik_aralik})",  RENK["para_format"]),
        (f"%{komisyon_yuzdesi:.0f} Komisyon (₺)",
                                   f"=H{bos+3}*{komisyon_yuzdesi/100}", RENK["para_format"]),
    ]

    for i, (etiket, deger, fmt) in enumerate(ozet_veriler):
        row = bos + 1 + i
        ws.merge_cells(f"A{row}:G{row}")
        et_hucre = ws[f"A{row}"]
        et_hucre.value     = etiket
        et_hucre.font      = Font(bold=True, name="Arial", size=10,
                                  color=RENK["baslik_yazi"])
        et_hucre.fill      = PatternFill("solid", fgColor=RENK["toplam_bg"])
        et_hucre.alignment = Alignment(horizontal="right", vertical="center")
        et_hucre.border    = _ince_cerceve()

        dg_hucre = ws.cell(row=row, column=9, value=deger)
        dg_hucre.font      = Font(bold=True, name="Arial", size=10,
                                  color=RENK["baslik_yazi"])
        dg_hucre.fill      = PatternFill("solid", fgColor=RENK["toplam_bg"])
        dg_hucre.alignment = Alignment(horizontal="right", vertical="center")
        dg_hucre.border    = _ince_cerceve()
        if fmt:
            dg_hucre.number_format = fmt

    # Satırları dondur (başlık sabit)
    ws.freeze_panes = "A4"


def _ozet_sheet_olustur(wb, satirlar, gun_str, komisyon_yuzdesi):
    ws = wb.create_sheet(title="ÖZET", index=0)

    ws.merge_cells("A1:F1")
    ws["A1"] = f"🏇 TJK Günlük Özet Raporu  —  {gun_str}"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor=RENK["baslik_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    basliklar = ["At Sahibi", "Koşuya Çıkan At", "Birinci",
                 "Toplam İkramiye (₺)", f"%{komisyon_yuzdesi:.0f} Komisyon (₺)", "Durum"]
    genislikler = [28, 16, 10, 22, 22, 12]

    for col_idx, (b, g) in enumerate(zip(basliklar, genislikler), 1):
        h = ws.cell(row=2, column=col_idx, value=b)
        h.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        h.fill      = PatternFill("solid", fgColor=RENK["altyazi_bg"])
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border    = _ince_cerceve()
        ws.column_dimensions[get_column_letter(col_idx)].width = g
    ws.row_dimensions[2].height = 22

    for row_idx, s in enumerate(satirlar, start=3):
        bg = RENK["satir_tek"] if row_idx % 2 == 0 else RENK["satir_cift"]
        durum = "✅ Kazandı" if s["birinci"] > 0 else "—"

        degerler = [
            s["sahip"], s["at_sayisi"], s["birinci"],
            s["toplam"], s["komisyon"], durum
        ]
        for col_idx, val in enumerate(degerler, 1):
            h = ws.cell(row=row_idx, column=col_idx, value=val)
            h.fill      = PatternFill("solid", fgColor=bg)
            h.border    = _ince_cerceve()
            h.font      = Font(name="Arial", size=10)
            h.alignment = Alignment(vertical="center")
            if col_idx in (4, 5):
                h.number_format = RENK["para_format"]
                h.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (2, 3, 6):
                h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    # Toplam satırı
    son = 2 + len(satirlar)
    toplam_row = son + 1
    ws.merge_cells(f"A{toplam_row}:C{toplam_row}")
    ws[f"A{toplam_row}"].value = "GENEL TOPLAM"
    ws[f"A{toplam_row}"].font  = Font(bold=True, color="FFFFFF", name="Arial")
    ws[f"A{toplam_row}"].fill  = PatternFill("solid", fgColor=RENK["toplam_bg"])
    ws[f"A{toplam_row}"].alignment = Alignment(horizontal="right")
    ws[f"A{toplam_row}"].border = _ince_cerceve()

    for col_idx, col_harf in [(4, "D"), (5, "E")]:
        h = ws[f"{col_harf}{toplam_row}"]
        h.value        = f"=SUM({col_harf}3:{col_harf}{son})"
        h.number_format = RENK["para_format"]
        h.font         = Font(bold=True, color="FFFFFF", name="Arial")
        h.fill         = PatternFill("solid", fgColor=RENK["toplam_bg"])
        h.alignment    = Alignment(horizontal="right", vertical="center")
        h.border       = _ince_cerceve()

    ws.freeze_panes = "A3"
